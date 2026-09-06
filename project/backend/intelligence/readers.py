"""
Turning a file somebody emailed us into a rectangle.

Two jobs, and the second is the interesting one.

Reading bytes is mechanical: sniff an encoding, sniff a delimiter, hand back
rows. Every importer does that.

**Finding the header is not mechanical**, and it is where naive importers
quietly destroy data. A spreadsheet a person maintained by hand does not start
with its column names -- it starts with the company's name, a note about when it
was last updated, and a blank row for air. Take row 0 as the header and every
subsequent column is mapped by a label that was never a label, which fails in
the worst possible way: silently, plausibly, and only visible three screens
later when somebody's salary is in the PAN field.

So the header is *found*, by scoring every candidate row against what headers
look like and data does not, and what was skipped is reported rather than
discarded. The number this file cares about most is `junk_rows_above`, because
telling the operator "your header is on line 4, I ignored three rows above it"
is the difference between a tool they trust and one they check by hand anyway.
"""

import csv
import io
import re
from dataclasses import dataclass, field

# Values that look like data rather than a column name. A row containing any of
# these is almost certainly a record, so it is penalised as a header candidate.
_DATA_SHAPES = [
    re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I),          # email
    re.compile(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$"),              # date
    re.compile(r"^\d{4}-\d{2}-\d{2}"),                           # iso date
    re.compile(r"^(rs\.?|inr)\s*[\d,]+", re.I),                  # money
    re.compile(r"^\+?\d[\d\s-]{8,}$"),                           # phone
    re.compile(r"^[A-Z]{5}\d{4}[A-Z]$"),                         # PAN
]

_TOTAL_ROW = re.compile(r"^\s*(grand\s+)?total\b|^\s*sum\s*$", re.I)

MAX_HEADER_SCAN = 15


@dataclass
class ParsedTable:
    headers: list
    rows: list
    header_row_index: int = 0
    raw_rows: list = field(default_factory=list)
    junk_rows_above: int = 0
    encoding: str = "utf-8"
    delimiter: str = ","
    sheet_name: str = None
    notes: list = field(default_factory=list)

    @property
    def row_count(self):
        return len(self.rows)

    @property
    def column_count(self):
        return len(self.headers)


# ==========================================================================
# Decoding
# ==========================================================================

#: Order matters. utf-8-sig first because Excel writes a BOM and leaving it on
#: turns the first header into "﻿Emp Naam", which then matches nothing.
#: cp1252 before latin-1 because it is what Windows actually produces, and
#: latin-1 last because it never raises -- it is the backstop, not a guess.
_ENCODINGS = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]


def decode_bytes(raw):
    for enc in _ENCODINGS:
        try:
            return raw.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1"


def sniff_delimiter(text):
    """
    Pick the delimiter that produces the most *consistent* column count.

    csv.Sniffer guesses from character frequency and is fooled by a file full
    of prose in one column, or by addresses containing commas under a
    semicolon-delimited export. Consistency is the better signal: the right
    delimiter makes every row the same width, and a wrong one does not.
    """
    sample = [ln for ln in text.split("\n")[:40] if ln.strip()]
    if not sample:
        return ","

    best, best_score = ",", -1.0
    for delim in [",", ";", "\t", "|"]:
        widths = [len(next(csv.reader([ln], delimiter=delim), [])) for ln in sample]
        widths = [w for w in widths if w > 0]
        if not widths:
            continue
        modal = max(set(widths), key=widths.count)
        if modal < 2:
            continue
        agreement = widths.count(modal) / len(widths)
        # Width breaks ties: two delimiters can both be perfectly consistent
        # when one of them simply never appears, and the one that actually
        # splits the file into columns is the real one.
        score = agreement * 100 + modal
        if score > best_score:
            best, best_score = delim, score
    return best


# ==========================================================================
# Header detection
# ==========================================================================

def _looks_like_data(cell):
    c = (cell or "").strip()
    return bool(c) and any(p.match(c) for p in _DATA_SHAPES)


def score_header_row(row, following):
    """
    How much does `row` look like a header, given the rows beneath it?

    Returns a float; higher is more header-like. The components are each worth
    stating because the UI shows the winner and an operator may disagree with
    it, and "it scored highest" is not an answer a person can argue with.
    """
    cells = [(c or "").strip() for c in row]
    filled = [c for c in cells if c]
    if len(filled) < 2:
        return 0.0

    score = 0.0

    # 1. A header labels every column it has. A title row fills one cell.
    fill_ratio = len(filled) / max(len(cells), 1)
    score += fill_ratio * 30

    # 2. Labels are distinct. Data repeats; column names do not.
    lowered = [c.lower() for c in filled]
    score += (len(set(lowered)) / len(lowered)) * 20

    # 3. Labels are short. "Date of joining" is a header; an address is not.
    avg_len = sum(len(c) for c in filled) / len(filled)
    score += 12 if avg_len <= 24 else (4 if avg_len <= 40 else -8)

    # 4. Labels are words, not values.
    data_like = sum(1 for c in filled if _looks_like_data(c))
    score -= (data_like / len(filled)) * 45

    # 5. Labels are rarely bare numbers.
    numeric = sum(1 for c in filled if re.match(r"^[\d.,]+$", c))
    score -= (numeric / len(filled)) * 30

    # 6. The rows below it should be as wide as it is. This is what separates
    #    a real header from a stray note that happens to have three words in
    #    it: a header predicts the shape of what follows.
    if following:
        widths = [len([c for c in r if (c or "").strip()]) for r in following[:8]]
        widths = [w for w in widths if w]
        if widths:
            avg_w = sum(widths) / len(widths)
            score += 18 * (1 - min(abs(avg_w - len(filled)) / max(len(filled), 1), 1.0))

    # 7. Earlier is better, all else equal -- a tie should not be broken by
    #    picking a row deep inside the data.
    return score


def detect_header(raw_rows):
    """Return (index, score) of the most header-like row in the first 15."""
    best_i, best_score = 0, float("-inf")
    limit = min(len(raw_rows), MAX_HEADER_SCAN)
    for i in range(limit):
        if not any((c or "").strip() for c in raw_rows[i]):
            continue
        score = score_header_row(raw_rows[i], raw_rows[i + 1:i + 9])
        score -= i * 0.5
        if score > best_score:
            best_i, best_score = i, score
    return best_i, best_score


# ==========================================================================
# Cleaning
# ==========================================================================

def _label_columns(header_row, width):
    """Name every column, including the ones the sheet forgot to name."""
    out = []
    seen = {}
    for i in range(width):
        raw = (header_row[i] if i < len(header_row) else "") or ""
        label = re.sub(r"\s+", " ", raw).strip()
        if not label:
            label = "Column %d" % (i + 1)
        # A duplicated header is not an error worth refusing over, but two
        # columns with the same name cannot both be mapped, so they are made
        # distinguishable here rather than fighting downstream.
        if label.lower() in seen:
            seen[label.lower()] += 1
            label = "%s (%d)" % (label, seen[label.lower()])
        else:
            seen[label.lower()] = 1
        out.append(label)
    return out


def _tidy(raw_rows, header_index):
    """Drop the noise a hand-kept sheet accumulates, and say what was dropped."""
    notes = []
    header_row = raw_rows[header_index]
    body = raw_rows[header_index + 1:]

    width = max([len([c for c in header_row if (c or "").strip()])] +
                [len(r) for r in body[:20]] or [len(header_row)])
    width = max(width, len(header_row))

    # Trailing summary row. It is not a person, and importing it as one creates
    # an employee called TOTAL earning eleven lakh a month.
    while body and _TOTAL_ROW.match((body[-1][0] if body[-1] else "") or ""):
        body.pop()
        notes.append("Ignored a trailing total row.")

    kept = []
    blank_dropped = 0
    for row in body:
        padded = list(row) + [""] * (width - len(row))
        if not any((c or "").strip() for c in padded):
            blank_dropped += 1
            continue
        kept.append([("" if c is None else str(c)) for c in padded[:width]])
    if blank_dropped:
        notes.append("Ignored %d blank row%s."
                     % (blank_dropped, "" if blank_dropped == 1 else "s"))

    headers = _label_columns(header_row, width)

    # Columns that are empty top to bottom carry nothing and only make the grid
    # harder to read, so they go -- but they are counted, because "you gave me
    # 14 columns and 3 were empty" is information.
    live = [i for i in range(width)
            if any((r[i] or "").strip() for r in kept) or headers[i].strip()]
    if len(live) < width:
        empty = width - len(live)
        notes.append("Ignored %d empty column%s."
                     % (empty, "" if empty == 1 else "s"))
        headers = [headers[i] for i in live]
        kept = [[r[i] for i in live] for r in kept]

    return headers, kept, notes


# ==========================================================================
# Entry points
# ==========================================================================

def _read_csv(raw, filename):
    text, encoding = decode_bytes(raw)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    delim = "\t" if filename.lower().endswith((".tsv", ".tab")) else sniff_delimiter(text)
    raw_rows = [r for r in csv.reader(io.StringIO(text), delimiter=delim)]
    return raw_rows, encoding, delim, None


def _read_xlsx(raw, filename):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_rows = []
    for row in ws.iter_rows(values_only=True):
        out = []
        for cell in row:
            if cell is None:
                out.append("")
            elif hasattr(cell, "isoformat"):
                # A real date cell is already unambiguous; keeping it ISO means
                # the profiler sees one format instead of inventing one.
                out.append(cell.isoformat()[:10])
            else:
                out.append(str(cell))
        raw_rows.append(out)
    sheet = ws.title
    wb.close()
    return raw_rows, "xlsx", None, sheet


def read_table(raw_bytes, filename):
    """Parse `raw_bytes` into a ParsedTable. `filename` chooses the reader."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        raw_rows, encoding, delim, sheet = _read_xlsx(raw_bytes, name)
    else:
        raw_rows, encoding, delim, sheet = _read_csv(raw_bytes, name)

    raw_rows = [list(r) for r in raw_rows]
    if not raw_rows:
        return ParsedTable(headers=[], rows=[], raw_rows=[],
                           notes=["The file is empty."])

    header_index, _ = detect_header(raw_rows)
    headers, rows, notes = _tidy(raw_rows, header_index)

    if header_index > 0:
        notes.insert(0, "Header found on line %d; ignored %d row%s above it."
                     % (header_index + 1, header_index,
                        "" if header_index == 1 else "s"))

    return ParsedTable(
        headers=headers, rows=rows, header_row_index=header_index,
        raw_rows=raw_rows, junk_rows_above=header_index, encoding=encoding,
        delimiter=delim, sheet_name=sheet, notes=notes)
