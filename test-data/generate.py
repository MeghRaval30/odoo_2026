"""
Build the rosters the import studio is demonstrated against.

Run from the repository root:

    python test-data/generate.py

Each file is a different *kind* of wrong, because the point of the demo is not
"we can read a spreadsheet". It is that the ways a real company's data actually
arrives are each handled by a different part of the pipeline.

  01  complete and tidy     nothing to fix. The control case.
  02  hand-kept             structural mess: junk rows above the header,
                            Hinglish column names, three date formats, rupees
                            written three ways, a TOTAL row at the bottom.
  03  legacy export         semantic mess: immaculate structure and wrong
                            anyway -- salary is annual where we store monthly,
                            blanks are the literal string NULL.
  04  incomplete            the demo file. Has people and pay, and no email,
                            no bank details, no employee codes at all.
  04b bank details          the supplement for 04. What finance keeps in its
                            own spreadsheet, keyed by name and staff id.
  05  acquisition           another company's vocabulary, plus four people who
                            already work here.

Deterministic: the RNG is seeded, so regenerating produces the same files and
the demo script's figures stay true. ASCII only in the data and in the output --
the Windows console is cp1252 and a rupee sign kills the command.
"""

import csv
import os
import random

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "import")

#: Four people already on the seeded roster, by the email the seed builds
#: ("{first}@oxp.com"). File 05 reuses them so duplicate detection has
#: something real to find.
COLLISIONS = [
    ("John", "Dsouza", "john@oxp.com"),
    ("Priya", "Sharma", "priya@oxp.com"),
    ("Meera", "Iyer", "meera@oxp.com"),
    ("Billy", "Kyle", "billy@oxp.com"),
]

HEAD_FILL = PatternFill("solid", fgColor="EDE3DC")


def _write_sheet(path, sheet_name, rows, header_row=1, widths=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)

    if header_row:
        for cell in ws[header_row]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = HEAD_FILL
                cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for i, width in enumerate(widths or [], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    wb.save(path)


def _pan(rng):
    return "%s%04d%s" % (
        "".join(rng.choice("ABCDEFGHIJKLMNOPQRSTUVWXYZ") for _ in range(5)),
        rng.randint(1000, 9999),
        rng.choice("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))


def _phone(rng):
    return "%d%d" % (rng.randint(70, 99), rng.randint(10000000, 99999999))


def _ifsc(rng):
    return rng.choice(["HDFC0000234", "ICIC0001177", "SBIN0007865",
                       "UTIB0000456", "KKBK0000958", "AXIS0000077"])


# ==========================================================================
# 01 -- complete and tidy
# ==========================================================================

CLEAN_PEOPLE = [
    ("Aditya", "Ranganathan", "Engineering", "Senior Developer", "2019-04-15", 118000),
    ("Nisha", "Verghese", "Engineering", "Developer", "2021-08-02", 78000),
    ("Rohan", "Bakshi", "Engineering", "Developer", "2022-01-17", 72000),
    ("Sneha", "Kulkarni", "Engineering", "QA Engineer", "2020-11-09", 68000),
    ("Vivek", "Nambiar", "Engineering", "Senior Developer", "2018-06-25", 125000),
    ("Ayesha", "Siddiqui", "Sales", "Account Manager", "2020-03-30", 92000),
    ("Kabir", "Chandra", "Sales", "Sales Executive", "2022-07-11", 61000),
    ("Tara", "Mukherjee", "Sales", "Sales Executive", "2021-05-24", 64000),
    ("Rishi", "Anand", "Sales", "Account Manager", "2019-09-16", 96000),
    ("Naina", "Kapadia", "Finance", "Accountant", "2020-02-10", 74000),
    ("Aryan", "Bhalla", "Finance", "Payroll Specialist", "2021-10-04", 81000),
    ("Ishani", "Roy", "Finance", "Accountant", "2022-04-19", 70000),
    ("Zoya", "Hussain", "HR", "HR Executive", "2021-01-11", 58000),
    ("Manav", "Grover", "HR", "Recruiter", "2022-09-05", 54000),
    ("Ritu", "Saxena", "HR", "HR Executive", "2020-08-21", 62000),
    ("Aakash", "Pandit", "IT", "Developer", "2019-12-02", 88000),
    ("Diya", "Menon", "IT", "QA Engineer", "2021-06-14", 71000),
    ("Nikhil", "Bhardwaj", "IT", "Developer", "2022-11-28", 76000),
    ("Simran", "Ahluwalia", "Engineering", "Developer", "2023-02-06", 69000),
    ("Om", "Prakash", "Engineering", "QA Engineer", "2023-05-22", 66000),
    ("Leela", "Krishnan", "Sales", "Sales Executive", "2023-08-14", 57000),
    ("Farhan", "Mirza", "Finance", "Accountant", "2023-03-27", 67000),
]


def build_complete(path):
    rng = random.Random(2601)
    rows = [["Employee Code", "First Name", "Last Name", "Work Email",
             "Department", "Job Position", "Date of Joining", "Monthly Salary",
             "Mobile Number", "Bank Account Number", "IFSC Code", "PAN"]]

    for i, (first, last, dept, role, doj, wage) in enumerate(CLEAN_PEOPLE, start=1):
        rows.append([
            "MER%04d" % (1000 + i), first, last,
            "%s.%s@meridiansystems.in" % (first.lower(), last.lower()),
            dept, role, doj, wage, _phone(rng),
            "%d" % rng.randint(10 ** 11, 10 ** 12 - 1), _ifsc(rng), _pan(rng),
        ])

    _write_sheet(path, "Employees", rows, header_row=1,
                 widths=[15, 14, 16, 34, 15, 20, 16, 15, 15, 20, 14, 13])
    return len(CLEAN_PEOPLE)


# ==========================================================================
# 02 -- the hand-kept spreadsheet
# ==========================================================================

MESSY_PEOPLE = [
    ("rajesh kumar",     "Engg", "Developer",          "15-03-2021", "Rs 45,000"),
    ("PRIYA NAIR",       "Engg", "Senior Developer",   "02/07/2019", "72,000"),
    ("  Anil Deshpande", "Sls",  "Sales Executive",    "2020-11-30", "38500/-"),
    ("Sneha Rao",        "Mktg", "Marketing Lead",     "11-01-2022", "Rs 65,000"),
    ("vikram Singh",     "Engg", "QA Engineer",        "23/08/2021", "Rs. 51,000"),
    ("Fatima Sheikh",    "HR",   "HR Executive",       "05-05-2020", "42,000"),
    ("Arjun MENON",      "Ops",  "Operations Analyst", "18/02/2022", "47500"),
    ("Kavya Reddy",      "Engg", "Developer",          "2021-06-14", "Rs 58,000"),
    ("  ROHIT Verma",    "Sls",  "Account Manager",    "09-09-2019", "68,000/-"),
    ("Divya Pillai",     "Mktg", "Content Writer",     "30/03/2023", "35,000"),
    ("Sameer Joshi",     "Engg", "Developer",          "12-12-2020", "Rs 54,500"),
    ("neha Bhatt",       "HR",   "Recruiter",          "07/07/2021", "40,000"),
    ("Imran Qureshi",    "Ops",  "Logistics Officer",  "2022-04-19", "44,000"),
    ("Pooja Agarwal",    "Sls",  "Sales Executive",    "25-10-2022", "36,500"),
    ("Karthik Nair",     "Engg", "Senior Developer",   "03/03/2018", "Rs 88,000"),
    ("ANJALI Desai",     "Mktg", "Marketing Lead",     "16-06-2021", "62,000"),
    ("Suresh Babu",      "Ops",  "Operations Analyst", "2019-08-08", "49,000"),
    ("Ritu Malhotra",    "HR",   "HR Executive",       "21/11/2022", "41,500"),
    ("Gaurav Sinha",     "Engg", "Developer",          "14-02-2023", "Rs 52,000"),
    ("meena Krishnan",   "Sls",  "Sales Executive",    "28/05/2020", "39,000"),
    ("Tarun Bhandari",   "Engg", "QA Engineer",        "2021-09-27", "50,500"),
    ("Lakshmi Iyer",     "Mktg", "Content Writer",     "06-06-2022", "Rs 37,000"),
]


def build_handmade(path):
    rng = random.Random(360)
    rows = [
        ["Brightloom Textiles Pvt Ltd - Employee Master", "", "", "", "",
         "", "", "", "", "", ""],
        ["Updated till 31 Aug 2026 (internal use only)", "", "", "", "",
         "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["Emp Naam", "Dept.", "Designation", "DOJ", "Sal (pm)", "Mob No",
         "Email ID", "A/C No", "IFSC", "PAN", "Remarks"],
    ]

    remarks = ["", "good performer", "", "on notice", "", "", "confirmed", "",
               "", "probation", "", "", "", "", "", "", "", "", "", "", "", ""]

    for i, (name, dept, role, doj, wage) in enumerate(MESSY_PEOPLE):
        phone = ("+91 " + _phone(rng)) if i % 3 == 0 else _phone(rng)

        email = "%s@brightloom.in" % name.strip().split()[0].lower()
        if i in (6, 17):
            email = ""                       # no address at all
        if i == 20:
            email = "rajesh@brightloom.in"   # a straight duplicate of row 1

        has_bank = i not in (3, 11)
        rows.append([
            name, dept, role, doj, wage, phone, email,
            "%d" % rng.randint(10 ** 11, 10 ** 12 - 1) if has_bank else "",
            _ifsc(rng) if has_bank else "",
            _pan(rng), remarks[i],
        ])

    # The summary row every hand-kept sheet grows. It is not a person, and
    # importing it as one creates an employee called TOTAL earning 11 lakh.
    rows.append(["TOTAL", "", "", "", "11 52 500", "", "", "", "", "", ""])

    _write_sheet(path, "Sheet1", rows, header_row=4,
                 widths=[20, 8, 20, 13, 13, 17, 26, 16, 14, 13, 16])

    wb = load_workbook(path)
    ws = wb.active
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"].font = Font(italic=True, size=9)
    wb.save(path)
    return len(MESSY_PEOPLE)


# ==========================================================================
# 03 -- the legacy export
# ==========================================================================

LEGACY_PEOPLE = [
    ("EMP0041", "Nikhil", "Raghavan", "Technology", "Software Engineer", "2020-02-17", 1080000),
    ("EMP0042", "Shreya", "Kulkarni", "Technology", "Senior Engineer", "2018-07-09", 1560000),
    ("EMP0043", "Manish", "Tiwari", "Finance", "Financial Analyst", "2021-01-25", 912000),
    ("EMP0044", "Aditi", "Chauhan", "People", "HR Business Partner", "2019-11-04", 1044000),
    ("EMP0045", "Sanjay", "Pillai", "Technology", "Engineering Manager", "2017-05-22", 2160000),
    ("EMP0046", "Ruchi", "Sethi", "Revenue", "Account Executive", "2022-03-14", 780000),
    ("EMP0047", "Ajay", "Kulkarni", "Operations", "Operations Lead", "2020-09-30", 996000),
    ("EMP0048", "Bhavna", "Mistry", "Finance", "Accounts Payable", "2021-08-16", 636000),
    ("EMP0049", "Deepak", "Choudhury", "Technology", "Software Engineer", "2022-06-06", 1020000),
    ("EMP0050", "Nandita", "Ghosh", "People", "Recruiter", "2023-02-13", 660000),
    ("EMP0051", "Pranav", "Salunkhe", "Revenue", "Account Executive", "2021-04-12", 828000),
    ("EMP0052", "Swati", "Chandran", "Technology", "QA Analyst", "2020-12-01", 744000),
    ("EMP0053", "Harish", "Balan", "Operations", "Supply Analyst", "2019-03-18", 888000),
    ("EMP0054", "Tanvi", "Mahajan", "Finance", "Financial Analyst", "2022-10-24", 864000),
    ("EMP0055", "Rohan", "Dixit", "Technology", "Senior Engineer", "2018-01-08", 1680000),
    ("EMP0056", "Ipsita", "Panda", "People", "HR Executive", "2023-05-29", 588000),
    ("EMP0057", "Varun", "Acharya", "Revenue", "Sales Manager", "2019-09-11", 1320000),
    ("EMP0058", "Kiran", "Hegde", "Operations", "Logistics Analyst", "2021-11-19", 792000),
]


def build_legacy(path):
    rng = random.Random(361)
    rows = [["EMPLOYEE_CODE", "FIRST_NAME", "LAST_NAME", "EMAIL_ADDRESS",
             "DEPARTMENT_NAME", "DESIGNATION", "DATE_OF_JOINING", "ANNUAL_CTC",
             "MOBILE_NUMBER", "BANK_AC", "IFSC_CODE", "PAN_NUMBER", "STATUS"]]

    for i, (code, first, last, dept, role, doj, ctc) in enumerate(LEGACY_PEOPLE):
        # A machine export writes NULL rather than leaving a cell empty, and
        # pads with trailing spaces. Both are silent poison to a naive import.
        blank = i in (2, 9)
        rows.append([
            code, first + "  ", last,
            "%s.%s@northgate-systems.com " % (first.lower(), last.lower()),
            dept, role, doj, ctc, _phone(rng),
            "NULL" if blank else "%d" % rng.randint(10 ** 11, 10 ** 12 - 1),
            "NULL" if blank else _ifsc(rng),
            _pan(rng), "Y" if i != 12 else "N",
        ])

    _write_sheet(path, "EMP_MASTER", rows, header_row=1,
                 widths=[16, 14, 14, 38, 18, 22, 18, 13, 16, 18, 14, 13, 9])
    return len(LEGACY_PEOPLE)


# ==========================================================================
# 04 -- incomplete, and 04b -- the supplement that completes it
# ==========================================================================

#: The demo file. It has people, roles and pay, and it is missing three things
#: the software needs: an email address, bank details, and an employee code.
#: Two people have no joining date and three have no salary, so a handful of
#: rows are genuinely unimportable and have to be reported rather than guessed.
#:
#: The staff id is here on purpose. It is the key the supplement joins on, and
#: it is deliberately NOT our employee code format -- the point of the demo is
#: that the operator asks the software to generate proper codes afterwards.
FIELDFORCE = [
    # staff_id, name, section, role, doj, wage
    ("FF-101", "Harpreet Sandhu",  "Field Ops", "Field Supervisor", "2021-03-08", 46000),
    ("FF-102", "Sunil Yadav",      "Field Ops", "Technician",       "2022-07-19", 32000),
    ("FF-103", "Rekha Dubey",      "Field Ops", "Technician",       "2021-11-02", 33500),
    ("FF-104", "Mohan Lal",        "Warehouse", "Storekeeper",      "2020-05-14", None),
    ("FF-105", "Geeta Rani",       "Warehouse", "Inventory Clerk",  "",           29000),
    ("FF-106", "Balwinder Singh",  "Field Ops", "Field Supervisor", "2019-09-23", 51000),
    ("FF-107", "Anita Kumari",     "Warehouse", "Storekeeper",      "2023-01-30", None),
    ("FF-108", "Ramesh Chandra",   "Field Ops", "Technician",       "2022-02-11", 31000),
    ("FF-109", "Kavita Devi",      "Warehouse", "Inventory Clerk",  "",           28500),
    ("FF-110", "Jaspal Gill",      "Field Ops", "Technician",       "2021-06-27", 34000),
    ("FF-111", "Shanti Bai",       "Warehouse", "Packer",           "2023-04-16", 26000),
    ("FF-112", "Vinod Kumar",      "Field Ops", "Field Supervisor", "2020-10-05", 48000),
    ("FF-113", "Pushpa Sharma",    "Warehouse", "Packer",           "2022-12-12", None),
    ("FF-114", "Deepak Rawat",     "Field Ops", "Technician",       "2023-08-21", 30500),
    ("FF-115", "Sarita Nair",      "Field Ops", "Technician",       "2022-05-30", 32500),
    ("FF-116", "Om Prakash Meena", "Warehouse", "Storekeeper",      "2021-02-18", 35000),
]


def build_incomplete(path):
    rng = random.Random(2604)
    rows = [["Staff ID", "Staff Name", "Section", "Role", "Joining Date",
             "Monthly Pay", "Contact"]]
    for staff_id, name, dept, role, doj, wage in FIELDFORCE:
        rows.append([staff_id, name, dept, role, doj,
                     "" if wage is None else wage, _phone(rng)])
    _write_sheet(path, "Staff", rows, header_row=1,
                 widths=[11, 22, 14, 20, 16, 14, 15])
    return len(FIELDFORCE)


def build_bank_supplement(path):
    """
    What finance keeps in its own spreadsheet.

    Deliberately not a perfect mirror of file 04. It is keyed by staff id and
    name, it is missing two of the sixteen people, and it carries one person
    who left before the migration -- which is exactly what a second file from
    another department looks like, and it gives the join something real to
    report: matched, not found, and unused.
    """
    rng = random.Random(2605)
    rows = [["Staff ID", "Employee Name", "Bank A/C Number", "IFSC",
             "PAN Number", "Account Type"]]

    missing = {"FF-107", "FF-113"}           # finance never received these two
    for staff_id, name, _dept, _role, _doj, _wage in FIELDFORCE:
        if staff_id in missing:
            continue
        rows.append([staff_id, name,
                     "%d" % rng.randint(10 ** 11, 10 ** 12 - 1),
                     _ifsc(rng), _pan(rng),
                     rng.choice(["Savings", "Savings", "Current"])])

    # Somebody who is on finance's list and not on HR's.
    rows.append(["FF-098", "Bhaskar Rao",
                 "%d" % rng.randint(10 ** 11, 10 ** 12 - 1),
                 _ifsc(rng), _pan(rng), "Savings"])

    _write_sheet(path, "Bank Details", rows, header_row=1,
                 widths=[11, 22, 20, 14, 14, 14])
    return len(rows) - 1


# ==========================================================================
# 05 -- the acquisition
# ==========================================================================

NORTHWIND_NEW = [
    ("Meghna", "Bharadwaj", "Technology", "Platform Engineer", "2021-07-05", 96000),
    ("Yash", "Trivedi", "Technology", "Data Engineer", "2022-02-21", 88000),
    ("Snehal", "Kadam", "Revenue", "Enterprise AE", "2020-10-12", 104000),
    ("Farhan", "Ali", "People Ops", "People Partner", "2021-12-06", 71000),
    ("Ritika", "Chopra", "Revenue", "Sales Development", "2023-01-30", 54000),
    ("Alok", "Nanda", "Technology", "Platform Engineer", "2019-06-17", 112000),
    ("Sunita", "Ravi", "People Ops", "Learning Lead", "2020-04-27", 79000),
    ("Devansh", "Kohli", "Technology", "Security Engineer", "2022-08-08", 99000),
]


def build_acquisition(path):
    rng = random.Random(362)
    rows = [["Employee Name", "Business Unit", "Role", "Start Date",
             "Monthly Pay (INR)", "Contact Number", "Work Email"]]

    for first, last, email in COLLISIONS:
        rows.append(["%s %s" % (first, last),
                     rng.choice(["Technology", "Revenue", "People Ops"]),
                     rng.choice(["Engineer", "Account Manager", "People Partner"]),
                     "2024-%02d-%02d" % (rng.randint(1, 12), rng.randint(1, 28)),
                     rng.randint(60, 120) * 1000, _phone(rng), email])

    for first, last, unit, role, start, pay in NORTHWIND_NEW:
        rows.append(["%s %s" % (first, last), unit, role, start, pay,
                     _phone(rng),
                     "%s.%s@northwind.co.in" % (first.lower(), last.lower())])

    # Kept as CSV on purpose: a due-diligence export arrives as a CSV far more
    # often than as a workbook, and it proves the reader is not xlsx-only.
    with open(path, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(rows)
    return len(rows) - 1


# ==========================================================================
# 06 -- a whole company, at the scale the product is actually used at
# ==========================================================================

#: Name pools rather than a literal list, because 240 hand-written rows would
#: be unreadable and would tempt somebody to trim it. Combined deterministically
#: under a fixed seed, so the file is byte-identical on every regeneration.
FIRST_NAMES = [
    "Aarav", "Aditi", "Advait", "Akshara", "Amrita", "Ananya", "Aniket",
    "Anjali", "Ankur", "Anusha", "Arnav", "Aruna", "Ashwin", "Avantika",
    "Bhavesh", "Chaitanya", "Charu", "Damini", "Darshan", "Deepika", "Dhruv",
    "Esha", "Gaurav", "Gayatri", "Girish", "Harini", "Hemant", "Indira",
    "Ishaan", "Jaya", "Kabir", "Kalpana", "Karan", "Keerthi", "Kunal",
    "Lavanya", "Madhav", "Mahesh", "Malini", "Manish", "Meghna", "Mihir",
    "Mohit", "Mrinal", "Naveen", "Neelam", "Nikhil", "Nithya", "Omkar",
    "Pallavi", "Parth", "Pooja", "Pranav", "Preeti", "Rahul", "Rajat",
    "Rakhi", "Ramya", "Rashmi", "Rohit", "Ruchira", "Sagar", "Sahana",
    "Sameer", "Sanjana", "Saurabh", "Shalini", "Shashank", "Shreya",
    "Siddharth", "Smita", "Sneha", "Sohail", "Sridhar", "Sunita", "Swapnil",
    "Tanvi", "Tarun", "Uday", "Ujjwal", "Vaishali", "Varun", "Vidya",
    "Vikas", "Vinay", "Vishal", "Yamini", "Yash", "Zara",
]

LAST_NAMES = [
    "Agarwal", "Ahuja", "Balakrishnan", "Banerjee", "Bhat", "Bhattacharya",
    "Chaudhary", "Chopra", "Deshmukh", "Dixit", "Dutta", "Gandhi", "Ghosh",
    "Gopalan", "Goswami", "Gupta", "Hegde", "Iyengar", "Jain", "Joshi",
    "Kamath", "Kaul", "Khanna", "Kohli", "Krishnan", "Kulkarni", "Lal",
    "Madhavan", "Malhotra", "Mehra", "Menon", "Mishra", "Mukherjee", "Nair",
    "Narayanan", "Patel", "Pillai", "Prasad", "Raghavan", "Rao", "Reddy",
    "Sahu", "Saxena", "Sengupta", "Shah", "Sharma", "Shetty", "Singh",
    "Sinha", "Srinivasan", "Subramanian", "Thakur", "Trivedi", "Varma",
    "Venkatesh", "Verma", "Wadhwa",
]

#: Department, role, and the band a wage is drawn from. Bands overlap the way
#: real ones do, and seniority is what moves somebody up one.
VANTAGE_ROLES = [
    ("Engineering", "Software Engineer",    62000,  95000, 34),
    ("Engineering", "Senior Engineer",     105000, 165000, 18),
    ("Engineering", "QA Engineer",          55000,  82000, 16),
    ("Engineering", "Engineering Manager", 175000, 240000,  6),
    ("Engineering", "DevOps Engineer",      78000, 125000, 10),
    ("Sales",       "Sales Executive",      42000,  68000, 22),
    ("Sales",       "Account Manager",      72000, 115000, 16),
    ("Sales",       "Sales Manager",       125000, 180000,  5),
    ("Finance",     "Accountant",           52000,  78000, 14),
    ("Finance",     "Payroll Specialist",   65000,  95000,  6),
    ("Finance",     "Financial Analyst",    82000, 125000,  8),
    ("HR",          "HR Executive",         45000,  70000, 12),
    ("HR",          "Recruiter",            48000,  75000,  9),
    ("HR",          "HR Business Partner",  95000, 140000,  4),
    ("IT",          "System Admin",         58000,  88000,  9),
    ("IT",          "Support Engineer",     40000,  62000, 15),
    ("Operations",  "Operations Analyst",   50000,  78000, 20),
    ("Operations",  "Logistics Officer",    44000,  66000, 16),
]

LOCATIONS = ["Mumbai", "Bengaluru", "Pune", "Hyderabad", "Remote"]


def build_bulk(path):
    """
    240 employees, in our own field names and clean throughout.

    The point of this file is not that it is hard to read -- it is that the
    import is worth doing at all. Twenty-two people is a demo; a company
    arrives with a few hundred, and the operation has to stay legible and fast
    at that size. It also gives every non-import screen -- the employee list,
    the dashboard, a payrun -- something the size of a real payroll to work on.
    """
    rng = random.Random(2606)

    roster = []
    for dept, role, low, high, count in VANTAGE_ROLES:
        for _ in range(count):
            roster.append((dept, role, low, high))
    rng.shuffle(roster)

    used_emails = set()
    rows = [["Employee Code", "First Name", "Last Name", "Work Email",
             "Department", "Job Position", "Work Location", "Employment Type",
             "Date of Joining", "Monthly Salary", "Mobile Number",
             "Bank Account Number", "IFSC Code", "PAN"]]

    for i, (dept, role, low, high) in enumerate(roster, start=1):
        first = rng.choice(FIRST_NAMES)
        last = rng.choice(LAST_NAMES)

        base = "%s.%s" % (first.lower(), last.lower())
        email = "%s@vantagelabs.in" % base
        n = 1
        while email in used_emails:
            n += 1
            email = "%s%d@vantagelabs.in" % (base, n)
        used_emails.add(email)

        # Joining dates spread over six years, weighted towards recent, which
        # is what a growing company's tenure distribution actually looks like.
        year = rng.choices([2019, 2020, 2021, 2022, 2023, 2024],
                           weights=[6, 9, 14, 20, 26, 25])[0]
        doj = "%d-%02d-%02d" % (year, rng.randint(1, 12), rng.randint(1, 28))

        etype = rng.choices(["Full Time", "Part Time", "Intern", "Contract"],
                            weights=[86, 5, 6, 3])[0]
        wage = rng.randrange(low, high, 500)
        if etype == "Intern":
            wage = rng.randrange(18000, 30000, 500)
        elif etype == "Part Time":
            wage = int(wage * 0.55 / 500) * 500

        rows.append([
            "VNT%04d" % (2000 + i), first, last, email, dept, role,
            rng.choice(LOCATIONS), etype, doj, wage, _phone(rng),
            "%d" % rng.randint(10 ** 11, 10 ** 12 - 1), _ifsc(rng), _pan(rng),
        ])

    _write_sheet(path, "Headcount", rows, header_row=1,
                 widths=[15, 14, 16, 32, 14, 22, 15, 16, 16, 15, 15, 20, 14, 13])
    return len(roster)




# ==========================================================================
# 07 -- three months of pay from the system this one replaces
# ==========================================================================

FILES = [
    ("01-meridian-complete.xlsx", build_complete),
    ("02-brightloom-handmade.xlsx", build_handmade),
    ("03-northgate-legacy-export.xlsx", build_legacy),
    ("04-fieldforce-incomplete.xlsx", build_incomplete),
    ("04b-fieldforce-bank-details.xlsx", build_bank_supplement),
    ("05-northwind-acquisition.csv", build_acquisition),
    ("06-vantage-240-headcount.xlsx", build_bulk),
]


def main():
    os.makedirs(HERE, exist_ok=True)
    for name, builder in FILES:
        count = builder(os.path.join(HERE, name))
        print("wrote %-36s %2d rows" % (name, count))
    print("")
    print("Open them from test-data/. See README.md for what each one breaks.")


if __name__ == "__main__":
    main()
